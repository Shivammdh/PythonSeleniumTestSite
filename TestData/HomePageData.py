import openpyxl


class HomePageData:
    test_Homepage_Data=[{'name':'Shivam Sharma','email':'shivamsharmamdh@gmail.com','gender':'Male'},
                            {'name':'Ayushi Chandra','email':'ayuhi27@gmail.com','gender':'Female'}]
    @staticmethod
    def getTestData(test_case_name):
        book = openpyxl.load_workbook(r"C:\Users\Asus\PycharmProjects\PythonSelFramework\TestData\testExcelData.xlsx")
        sheet = book.active
        Dict={}
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return[Dict]