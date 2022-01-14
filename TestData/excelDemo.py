import openpyxl
book=openpyxl.load_workbook("testExcelData.xlsx")
sheet=book.active
cell=sheet.cell(row=2,column=2)
Dict={}
print(cell.value)
print(sheet.max_row)
print(sheet.max_column)
for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i,column=1).value=="Testcase1":
        for j in range(2,sheet.max_column+1):
            Dict[sheet.cell(row=1,column=j).value]=sheet.cell(row=i,column=j).value





